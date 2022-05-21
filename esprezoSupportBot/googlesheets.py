
from pprint import pprint

import httplib2
import apiclient.discovery
from oauth2client.service_account import ServiceAccountCredentials

from openpyxl.utils.cell import get_column_letter

from config import SPREADSHEET_ID

# Файл, полученный в Google Developer Console
CREDENTIALS_FILE = '/home/vtyree/esprezoSupportBot/esprezoSupportBot/config/creds.json'
SCOPES = ['https://www.googleapis.com/auth/spreadsheets',
		'https://www.googleapis.com/auth/drive']

# Авторизуемся и получаем service — экземпляр доступа к API
def sheetautorize():
	credentials = ServiceAccountCredentials.from_json_keyfile_name(CREDENTIALS_FILE, SCOPES)
	httpAuth = credentials.authorize(httplib2.Http())
	return (apiclient.discovery.build('sheets', 'v4', http = httpAuth))

service = sheetautorize()

# Пример чтения файла
#values = service.spreadsheets().values().get(
#	spreadsheetId = SPREADSHEET_ID,
#	range = 'A1:E10',
#	majorDimension = 'COLUMNS'
#).execute()
#pprint(values)


def check_id_duplication(uid: str):
	response = service.spreadsheets().values().get(
		spreadsheetId = SPREADSHEET_ID,
		majorDimension = "COLUMNS",
		range = "main!A1:A"
	).execute()
	for i in response['values'][0]:
		if str(uid) == i:
			return(0)
	return(1)

def check_mail_duplication(mail: str):
	response = service.spreadsheets().values().get(
		spreadsheetId = SPREADSHEET_ID,
		majorDimension = "COLUMNS",
		range = "main!C1:C"
	).execute()
	for i in response['values'][0]:
		if mail == i:
			return(0)
	return(1)

def get_mails():
	response = service.spreadsheets().values().get(
		spreadsheetId = SPREADSHEET_ID,
		range = "letters!A2:A",
		majorDimension = "COLUMNS",
	).execute()
	mails = response['values'][0]
	return(mails)

def start_register(user_data: dict):
	response = service.spreadsheets().values().get(
		spreadsheetId = SPREADSHEET_ID,
		range = "main!A1:A"
	).execute()
	last_row = len(response['values']) + 1
	values = service.spreadsheets().values().batchUpdate(
		spreadsheetId = SPREADSHEET_ID,
		body={
			"valueInputOption": "USER_ENTERED",
			"data": [
				{"range": f"main!A{last_row}:C{last_row}",
				"majorDimension": "ROWS",
				"values": [[f"{user_data[0]}", f"{user_data[1]}",  f"{user_data[2]}"]]},
		]
	}
	).execute()
	pprint(values)

def check_if_registred(uid):
	response = service.spreadsheets().values().get(
		spreadsheetId = SPREADSHEET_ID,
		range = "main!A1:A",
		majorDimension = "COLUMNS",
	).execute()
	id_row = response['values'][0].index(str(uid)) + 1
	response = service.spreadsheets().values().get(
		spreadsheetId = SPREADSHEET_ID,
		range = f"main!D{id_row}:G{id_row}",
		majorDimension = "ROWS",
	).execute()
	try:
		data = response['values'][0]
	except:
		return (0)
	return (data)

def checkup_register(uid, user_data: dict):
	response = service.spreadsheets().values().get(
		spreadsheetId = SPREADSHEET_ID,
		range = "main!A1:A",
		majorDimension = "COLUMNS",
	).execute()
	id_row = response['values'][0].index(str(uid)) + 1
	if (len(user_data) == 4):
		values = service.spreadsheets().values().batchUpdate(
			spreadsheetId = SPREADSHEET_ID,
			body={
				"valueInputOption": "USER_ENTERED",
				"data": [
					{"range": f"main!D{id_row}:G{id_row}",
					"majorDimension": "ROWS",
					"values": [[f"{user_data[0]}", f"{user_data[1]}", f"MSK {user_data[2]}", f"{user_data[3]}"]]},
			]
		}
		).execute()
	else:
		values = service.spreadsheets().values().batchUpdate(
			spreadsheetId = SPREADSHEET_ID,
			body={
				"valueInputOption": "USER_ENTERED",
				"data": [
					{"range": f"main!F{id_row}:G{id_row}",
					"majorDimension": "ROWS",
					"values": [[f"MSK {user_data[2]}", f"{user_data[3]}"]]},
			]
		}
		).execute()
	pprint(values)

def get_name_by_id(uid: str):
	response = service.spreadsheets().values().get(
		spreadsheetId = SPREADSHEET_ID,
		range = "main!A1:A",
		majorDimension = "COLUMNS",
	).execute()
	id_row = response['values'][0].index(str(uid)) + 1
	response = service.spreadsheets().values().get(
		spreadsheetId = SPREADSHEET_ID,
		range = f"main!B{id_row}",
		majorDimension = "COLUMNS",
	).execute()
	name = response["values"][0][0]
	return (name)

def checkup_freq_update(uid, user_data: dict):
	response = service.spreadsheets().values().get(
		spreadsheetId = SPREADSHEET_ID,
		range = "main!A1:A",
		majorDimension = "COLUMNS",
	).execute()
	id_row = response['values'][0].index(str(uid)) + 1
	values = service.spreadsheets().values().batchUpdate(
		spreadsheetId = SPREADSHEET_ID,
		body={
			"valueInputOption": "USER_ENTERED",
			"data": [
				{"range": f"main!D{id_row}:E{id_row}",
				"majorDimension": "ROWS",
				"values": [[f"{user_data[0]}", f"{user_data[1]}"]]},
		]
	}
	).execute()

def update_name(uid, name: str):
	response = service.spreadsheets().values().get(
		spreadsheetId = SPREADSHEET_ID,
		range = "main!A1:A",
		majorDimension = "COLUMNS",
	).execute()
	id_row = response['values'][0].index(str(uid)) + 1
	values = service.spreadsheets().values().batchUpdate(
		spreadsheetId = SPREADSHEET_ID,
		body={
			"valueInputOption": "USER_ENTERED",
			"data": [
				{"range": f"main!B{id_row}",
				"majorDimension": "ROWS",
				"values": [[f"{name}"]]},
		]
	}
	).execute()

def update_mail(uid, mail: str):
	response = service.spreadsheets().values().get(
		spreadsheetId = SPREADSHEET_ID,
		range = "main!A1:A",
		majorDimension = "COLUMNS",
	).execute()
	id_row = response['values'][0].index(str(uid)) + 1
	values = service.spreadsheets().values().batchUpdate(
		spreadsheetId = SPREADSHEET_ID,
		body={
			"valueInputOption": "USER_ENTERED",
			"data": [
				{"range": f"main!C{id_row}",
				"majorDimension": "ROWS",
				"values": [[f"{mail}"]]},
		]
	}
	).execute()


def get_letters(uid):
	response = service.spreadsheets().values().get(
		spreadsheetId = SPREADSHEET_ID,
		range = "main!A1:A",
		majorDimension = "COLUMNS",
	).execute()
	id_row = response['values'][0].index(str(uid)) + 1
	response = service.spreadsheets().values().get(
		spreadsheetId = SPREADSHEET_ID,
		range = f"main!C{id_row}",
		majorDimension = "COLUMNS",
	).execute()
	mail = response['values'][0][0]
	response = service.spreadsheets().values().get(
		spreadsheetId = SPREADSHEET_ID,
		range = f"letters!A1:A",
		majorDimension = "COLUMNS",
	).execute()
	mail_row = response['values'][0].index(str(mail)) + 1
	response = service.spreadsheets().values().get(
		spreadsheetId = SPREADSHEET_ID,
		range = f"letters!C{mail_row}:{mail_row}",
		majorDimension = "ROWS",
	).execute()
	try:
		letters = response["values"][0]
	except:
		letters = 0
	return(letters)

def get_drafts(uid):
	response = service.spreadsheets().values().get(
		spreadsheetId = SPREADSHEET_ID,
		range = "drafts!A1:A",
		majorDimension = "COLUMNS",
	).execute()
	try:
		id_row = response['values'][0].index(str(uid)) + 1
		response = service.spreadsheets().values().get(
			spreadsheetId = SPREADSHEET_ID,
			range = f"drafts!B{id_row}:{id_row}",
			majorDimension = "ROWS",
		).execute()
		drafts = response["values"][0]
	except:
		drafts = 0
	return(drafts)

def get_notes(uid):
	response = service.spreadsheets().values().get(
		spreadsheetId = SPREADSHEET_ID,
		range = "checkup!A1:1",
		majorDimension = "ROWS",
	).execute()
	try:
		id_row = response['values'][0].index(str(uid)) + 1
		response = service.spreadsheets().values().get(
			spreadsheetId = SPREADSHEET_ID,
			range = f"checkup!B3:B",
			majorDimension = "COLUMNS",
		).execute()
		notes = response["values"][0]
	except:
		notes = 0
	return(notes)

def get_user_name(uid):
	response = service.spreadsheets().values().get(
		spreadsheetId = SPREADSHEET_ID,
		range = "main!A1:A",
		majorDimension = "COLUMNS",
	).execute()
	id_row = response['values'][0].index(str(uid)) + 1
	response = service.spreadsheets().values().get(
		spreadsheetId = SPREADSHEET_ID,
		range = f"main!C{id_row}",
		majorDimension = "COLUMNS",
	).execute()
	mail = response['values'][0][0]
	response = service.spreadsheets().values().get(
		spreadsheetId = SPREADSHEET_ID,
		range = f"letters!A1:A",
		majorDimension = "COLUMNS",
	).execute()
	mail_row = response['values'][0].index(str(mail)) + 1
	response = service.spreadsheets().values().get(
		spreadsheetId = SPREADSHEET_ID,
		range = f"letters!B{mail_row}",
		majorDimension = "ROWS",
	).execute()
	name = response["values"][0][0]
	return(name)

def get_user_id_from_grattitude_name(name):
	response = service.spreadsheets().values().get(
		spreadsheetId = SPREADSHEET_ID,
		range = "letters!B1:B",
		majorDimension = "COLUMNS",
	).execute()
	#print(response['values'][0][0])
	id_row = response['values'][0].index(str(name)) + 1
	response = service.spreadsheets().values().get(
		spreadsheetId = SPREADSHEET_ID,
		range = f"letters!A{id_row}",
		majorDimension = "COLUMNS",
	).execute()
	mail = response["values"][0][0]
	response = service.spreadsheets().values().get(
		spreadsheetId = SPREADSHEET_ID,
		range = f"main!C1:C",
		majorDimension = "COLUMNS",
	).execute()
	id_row = response['values'][0].index(str(mail)) + 1
	response = service.spreadsheets().values().get(
		spreadsheetId = SPREADSHEET_ID,
		range = f"main!A{id_row}",
		majorDimension = "COLUMNS",
	).execute()
	uid = response['values'][0][0]
	return(uid)

def save_message(name, message):
	response = service.spreadsheets().values().get(
		spreadsheetId = SPREADSHEET_ID,
		range = "letters!B:B",
		majorDimension = "COLUMNS",
	).execute()
	name_row = response['values'][0].index(str(name)) + 1
	response = service.spreadsheets().values().get(
		spreadsheetId = SPREADSHEET_ID,
		range = f"letters!C{name_row}:{name_row}",
		majorDimension = "ROWS",
	).execute()
	count_letters =  len(response['values'][0])
	service.spreadsheets().values().batchUpdate(
		spreadsheetId = SPREADSHEET_ID,
		body={
			"valueInputOption": "USER_ENTERED",
			"data": [
				{"range": f"letters!{get_column_letter(count_letters + 3)}{name_row}:{get_column_letter(count_letters + 3)}",
				"majorDimension": "ROWS",
				"values": [[message]]},
		]
	}
	).execute()

def delete_message(name):
	response = service.spreadsheets().values().get(
		spreadsheetId = SPREADSHEET_ID,
		range = "letters!B:B",
		majorDimension = "COLUMNS",
	).execute()
	name_row = response['values'][0].index(str(name)) + 1
	response = service.spreadsheets().values().get(
		spreadsheetId = SPREADSHEET_ID,
		range = f"letters!C{name_row}:{name_row}",
		majorDimension = "ROWS",
	).execute()
	count_letters =  len(response['values'][0])
	service.spreadsheets().values().batchUpdate(
		spreadsheetId = SPREADSHEET_ID,
		body={
			"valueInputOption": "USER_ENTERED",
			"data": [
				{"range": f"letters!{get_column_letter(count_letters + 2)}{name_row}:{get_column_letter(count_letters + 2)}",
				"majorDimension": "ROWS",
				"values": [[""]]},
		]
	}
	).execute()

def save_to_drafts(uid, data):
	response = service.spreadsheets().values().get(
		spreadsheetId = SPREADSHEET_ID,
		majorDimension = "COLUMNS",
		range = "drafts!A1:A"
	).execute()
	try:
		id_row = response['values'][0].index(str(uid)) + 1
		response = service.spreadsheets().values().get(
			spreadsheetId = SPREADSHEET_ID,
			range = f"drafts!B{id_row}:{id_row}",
			majorDimension = "ROWS",
		).execute()
		count_drafts =  len(response['values'][0])
		service.spreadsheets().values().batchUpdate(
		spreadsheetId = SPREADSHEET_ID,
		body={
			"valueInputOption": "USER_ENTERED",
			"data": [
				{"range": f"drafts!{get_column_letter(count_drafts + 2)}{id_row}:{get_column_letter(count_drafts + 2)}",
				"majorDimension": "ROWS",
				"values": [[str(data)]]},
			]
		}
		).execute()
	except:
		service.spreadsheets().values().append(
		spreadsheetId = SPREADSHEET_ID,
		valueInputOption = "USER_ENTERED",
		range = "drafts!A1",
		body={
				"majorDimension": "ROWS",
				"values": [[uid, str(data)]]},
		).execute()

def save_checkup_data(data, uid):
	response_id = service.spreadsheets().values().get(
		spreadsheetId = SPREADSHEET_ID,
		majorDimension = "ROWS",
		range = "checkup!A1:1"
	).execute()
	try:
		id_col = response_id['values'][0].index(str(uid)) + 1
		response = service.spreadsheets().values().get(
			spreadsheetId = SPREADSHEET_ID,
			range = f"checkup!{get_column_letter(id_col)}3:{get_column_letter(id_col)}",
			majorDimension = "COLUMNS",
		).execute()
		count_help =  len(response['values'][0])
		service.spreadsheets().values().batchUpdate(
		spreadsheetId = SPREADSHEET_ID,
		body={
			"valueInputOption": "USER_ENTERED",
			"data": [
				{"range": f"checkup!{get_column_letter(id_col)}{count_help + 3}:{get_column_letter(id_col)}{count_help + 3}",
				"majorDimension": "ROWS",
				"values": [[f"{data}"]],}
			]
		}
		).execute()
	except:
		id_col = len(response_id['values'][0])
		service.spreadsheets().values().batchUpdate(
		spreadsheetId = SPREADSHEET_ID,
		body={
			"valueInputOption": "USER_ENTERED",
			"data": [
				{"range": f"checkup!{get_column_letter(id_col + 1)}1:{get_column_letter(id_col + 1)}3",
				"majorDimension": "COLUMNS",
				"values": [[f"{uid}", f"{get_user_name(uid)}", f"{str(data)}"]]}
			]
		}
		).execute()

def save_help_data(data, uid):
	response_id = service.spreadsheets().values().get(
		spreadsheetId = SPREADSHEET_ID,
		majorDimension = "ROWS",
		range = "help!A1:1"
	).execute()
	try:
		id_col = response_id['values'][0].index(str(uid)) + 1
		response = service.spreadsheets().values().get(
			spreadsheetId = SPREADSHEET_ID,
			range = f"help!{get_column_letter(id_col)}3:{get_column_letter(id_col)}",
			majorDimension = "COLUMNS",
		).execute()
		count_help =  len(response['values'][0])
		service.spreadsheets().values().batchUpdate(
		spreadsheetId = SPREADSHEET_ID,
		body={
			"valueInputOption": "USER_ENTERED",
			"data": [
				{"range": f"help!{get_column_letter(id_col)}{count_help + 3}:{get_column_letter(id_col)}{count_help + 3}",
				"majorDimension": "ROWS",
				"values": [[f"{str(data)}"]],}
			]
		}
		).execute()
	except:
		id_col = len(response_id['values'][0])
		service.spreadsheets().values().batchUpdate(
		spreadsheetId = SPREADSHEET_ID,
		body={
			"valueInputOption": "USER_ENTERED",
			"data": [
				{"range": f"help!{get_column_letter(id_col + 1)}1:{get_column_letter(id_col + 1)}3",
				"majorDimension": "COLUMNS",
				"values": [[f"{uid}", f"{get_user_name(uid)}", f"{str(data)}"]]}
			]
		}
		).execute()



